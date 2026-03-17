"""
Blockchain Audit Verification System

A terminal-based demonstration tool showing how two parties to a transaction
can independently produce matching SHA-256 hashes through standardized data
entry, enabling real-time decentralized audit verification.
"""

import hashlib
import json
import os
import re
from datetime import datetime
from collections import defaultdict

from rich.console import Console
from rich.table import Table
from openpyxl import Workbook, load_workbook

console = Console()

# --- Constants ---

LEDGER_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "ledger.json")

COMPANY_KEYS = {
    "COMPANYA": {"role": "company", "entity_name": "COMPANYA"},
    "VENDORB":  {"role": "company", "entity_name": "VENDORB"},
}
AUDITOR_KEY = "VIEW"

FIELD_ORDER = [
    "transaction_id",
    "date",
    "product_id",
    "amount",
    "currency",
    "quantity",
    "payer_id",
    "seller_id",
    "on_credit",
]

# --- Utility Functions ---


def clear_screen():
    os.system("cls" if os.name == "nt" else "clear")


def display_banner():
    print("=" * 60)
    print("  Blockchain Transaction Verification System")
    print("  Company A  |  Vendor B")
    print(f"  {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("=" * 60)
    print()


def ask_choice(prompt, choices, default=None):
    """Prompt the user to choose from a list of options."""
    while True:
        suffix = f" [{'/'.join(choices)}]"
        if default:
            suffix += f" (default: {default})"
        val = input(f"{prompt}{suffix}: ").strip().lower()
        if val == "" and default is not None:
            return default
        if val in choices:
            return val
        print(f"  Please enter one of: {', '.join(choices)}")


# --- Validation Functions ---


def validate_date(value):
    """Parse and standardize a date to YYYY-MM-DD format."""
    value = value.strip()
    try:
        parsed = datetime.strptime(value, "%Y-%m-%d")
        return parsed.strftime("%Y-%m-%d")
    except ValueError:
        return None


def validate_amount(value):
    """Strip $ and commas, return formatted to 2 decimal places."""
    value = re.sub(r"[$,]", "", value.strip())
    try:
        num = float(value)
        if num < 0:
            return None
        return f"{num:.2f}"
    except ValueError:
        return None


def validate_quantity(value):
    """Parse a positive number, format to 2 decimal places."""
    value = value.strip()
    try:
        num = float(value)
        if num < 0:
            return None
        return f"{num:.2f}"
    except ValueError:
        return None


def validate_yes_no(value):
    """Normalize yes/no input to YES or NO."""
    value = value.strip().lower()
    if value in ("yes", "y", "true", "1"):
        return "YES"
    elif value in ("no", "n", "false", "0"):
        return "NO"
    return None


def validate_non_empty(value):
    """Check that the value is not empty after stripping."""
    value = value.strip()
    return value if value else None


def prompt_field(field_name, validator=None, hint=""):
    """Prompt the user for a field value with optional validation."""
    while True:
        suffix = f" ({hint})" if hint else ""
        value = input(f"  {field_name}{suffix}: ").strip()
        if validator:
            result = validator(value)
            if result is None:
                print("    Invalid input. Please try again.")
                continue
            return result
        else:
            if not value:
                print("    This field cannot be empty.")
                continue
            return value


# --- Normalization and Hashing ---


def normalize_transaction(fields):
    """
    Normalize raw input fields into a canonical pipe-delimited string.
    Returns the canonical string ready for hashing.
    """
    normalized = {}

    # Transaction ID: uppercase, strip spaces
    normalized["transaction_id"] = fields["transaction_id"].upper().replace(" ", "")

    # Date: already validated to YYYY-MM-DD
    normalized["date"] = fields["date"]

    # Product ID: uppercase, strip spaces
    normalized["product_id"] = fields["product_id"].upper().replace(" ", "")

    # Amount: already validated to 2 decimals
    normalized["amount"] = fields["amount"]

    # Currency: uppercase
    normalized["currency"] = fields["currency"].upper().strip()

    # Quantity: already validated to 2 decimals
    normalized["quantity"] = fields["quantity"]

    # Payer ID: uppercase, strip spaces
    normalized["payer_id"] = fields["payer_id"].upper().replace(" ", "")

    # Seller ID: uppercase, strip spaces
    normalized["seller_id"] = fields["seller_id"].upper().replace(" ", "")

    # On Credit: already validated to YES/NO
    normalized["on_credit"] = fields["on_credit"]

    # Build canonical string in fixed field order
    canonical = "|".join(normalized[f] for f in FIELD_ORDER)
    return canonical, normalized


def hash_transaction(canonical_string):
    """Produce a SHA-256 hex digest of the canonical string."""
    return hashlib.sha256(canonical_string.encode("utf-8")).hexdigest()


# --- Ledger Persistence ---


def load_ledger():
    """Load the ledger from disk, or return an empty ledger."""
    try:
        with open(LEDGER_FILE, "r") as f:
            return json.load(f)
    except (FileNotFoundError, json.JSONDecodeError):
        return {"entries": []}


def save_ledger(ledger):
    """Write the ledger to disk."""
    with open(LEDGER_FILE, "w") as f:
        json.dump(ledger, f, indent=2)


def get_next_id(ledger):
    """Generate the next entry ID (ENTRY-001, ENTRY-002, etc.)."""
    count = len(ledger["entries"])
    return f"ENTRY-{count + 1:03d}"


# --- Display Functions ---


def display_transaction_summary(fields, normalized, canonical, hash_value):
    """Show the user their input, normalized values, canonical string, and hash."""
    table = Table(title="Transaction Summary", show_lines=True)
    table.add_column("Field", style="bold")
    table.add_column("Your Input")
    table.add_column("Normalized")

    field_labels = {
        "transaction_id": "Transaction ID",
        "date": "Date (UTC)",
        "product_id": "Product ID",
        "amount": "Total Amount",
        "currency": "Currency",
        "quantity": "Quantity",
        "payer_id": "Payer ID",
        "seller_id": "Seller ID",
        "on_credit": "On Credit?",
    }

    for key in FIELD_ORDER:
        table.add_row(field_labels[key], str(fields[key]), str(normalized[key]))

    console.print(table)
    print()
    print("Canonical String:")
    print(f"  {canonical}")
    print()
    print("SHA-256 Hash:")
    print(f"  {hash_value}")
    print()


# --- Company Flow ---


def company_flow(user):
    """Main workflow for Company A or Vendor B: enter transaction data, hash, and submit."""
    while True:
        clear_screen()
        display_banner()
        print(f"Logged in as: {user['entity_name']}")
        print()
        print("Enter Transaction Data")
        print("Fill in each field below. Data will be normalized before hashing.")
        print()

        # Collect raw inputs
        fields = {}
        fields["transaction_id"] = prompt_field(
            "Transaction ID", validate_non_empty, "e.g. INV-001"
        )
        fields["date"] = prompt_field(
            "Date (UTC)", validate_date, "YYYY-MM-DD"
        )
        fields["product_id"] = prompt_field(
            "Product ID", validate_non_empty, "e.g. WIDGET-500"
        )
        fields["amount"] = prompt_field(
            "Total Amount", validate_amount, "e.g. 10000 or $10,000"
        )
        fields["currency"] = prompt_field(
            "Currency", validate_non_empty, "e.g. USD"
        )
        fields["quantity"] = prompt_field(
            "Quantity", validate_quantity, "e.g. 100"
        )
        fields["payer_id"] = prompt_field(
            "Payer ID", validate_non_empty, "e.g. COMPANYA"
        )
        fields["seller_id"] = prompt_field(
            "Seller ID", validate_non_empty, "e.g. VENDORB"
        )
        fields["on_credit"] = prompt_field(
            "On Credit?", validate_yes_no, "yes/no"
        )

        # Normalize and hash
        canonical, normalized = normalize_transaction(fields)
        hash_value = hash_transaction(canonical)

        print()
        display_transaction_summary(fields, normalized, canonical, hash_value)

        # Confirm submission
        submit = ask_choice("Submit this transaction to the ledger?", ["y", "n"], default="y")
        if submit == "y":
            ledger = load_ledger()
            entry = {
                "id": get_next_id(ledger),
                "transaction_id": normalized["transaction_id"],
                "hash": hash_value,
                "submitter": user["entity_name"],
                "timestamp": datetime.utcnow().strftime("%Y-%m-%dT%H:%M:%S"),
            }
            ledger["entries"].append(entry)
            save_ledger(ledger)
            print(f"Transaction submitted as {entry['id']}.")
        else:
            print("Transaction discarded.")

        print()
        again = ask_choice("Enter another transaction?", ["y", "n"], default="n")
        if again != "y":
            break


# --- Auditor Flow ---


def find_matches(ledger):
    """
    Analyze the ledger and group entries purely by hash output.
    Entries sharing the same hash from different submitters are VERIFIED.
    Entries whose hash has no counterparty match are UNMATCHED and flagged.
    """
    # Group entries by their hash value
    hash_groups = defaultdict(list)
    for entry in ledger["entries"]:
        hash_groups[entry["hash"]].append(entry)

    matched = []
    unmatched = []

    for hash_val, entries in hash_groups.items():
        submitters = set(e["submitter"] for e in entries)
        if len(submitters) >= 2:
            # Multiple parties produced the same hash — verified
            matched.append({
                "hash": hash_val,
                "entries": entries,
                "status": "VERIFIED",
            })
        else:
            # Only one party submitted this hash — unmatched
            for entry in entries:
                unmatched.append({
                    "hash": hash_val,
                    "entries": [entry],
                    "status": "UNMATCHED",
                })

    # Sort matched groups by earliest timestamp, then unmatched
    matched.sort(key=lambda g: min(e["timestamp"] for e in g["entries"]))
    unmatched.sort(key=lambda g: g["entries"][0]["timestamp"])

    return matched + unmatched


def display_chain_stats(ledger):
    """Display a summary of blockchain ledger statistics."""
    matches = find_matches(ledger)
    total_txns = len(ledger["entries"])
    verified = sum(len(m["entries"]) for m in matches if m["status"] == "VERIFIED")
    flagged = sum(1 for m in matches if m["status"] == "UNMATCHED")
    print(
        f"  1 Block  |  {total_txns} Entries"
        f"  |  Verified: {verified}"
        f"  |  Flagged: {flagged}"
    )
    print()


def display_verification_status(ledger):
    """Show per-transaction hash verification status."""
    matches = find_matches(ledger)

    if not matches:
        print("No transactions in the ledger.")
        return matches

    display_chain_stats(ledger)

    verified = [m for m in matches if m["status"] == "VERIFIED"]
    unmatched = [m for m in matches if m["status"] == "UNMATCHED"]

    group_num = 1

    if verified:
        print("=== VERIFIED MATCHES ===")
        print()
        for match in verified:
            print(f"--- Entry Group {group_num} ---")
            print(f"  Status: VERIFIED")
            print(f"  Hash: {match['hash']}")
            print()
            for entry in match["entries"]:
                print(f"  {entry['submitter']}  {entry['timestamp']}  ({entry['id']})")
            print()
            print("  Hashes match -- Transaction verified")
            print()
            group_num += 1

    if unmatched:
        print("=== UNMATCHED -- FLAGGED FOR REVIEW ===")
        print()
        for match in unmatched:
            entry = match["entries"][0]
            print(f"--- Entry Group {group_num} ---")
            print(f"  Status: UNMATCHED -- FLAGGED")
            print(f"  Hash: {entry['hash']}")
            print()
            print(f"  {entry['submitter']}  {entry['timestamp']}  ({entry['id']})")
            print()
            print("  No counterparty hash match -- Flagged for review")
            print()
            group_num += 1

    return matches


def display_full_ledger(ledger):
    """Show all transactions in the ledger."""
    if not ledger["entries"]:
        print("No transactions in the ledger.")
        return

    display_chain_stats(ledger)

    first_ts = ledger["entries"][0]["timestamp"]
    last_ts = ledger["entries"][-1]["timestamp"]
    print(f"  Block #1")
    print(f"  Transactions: {len(ledger['entries'])}")
    print(f"  First entry:  {first_ts}")
    print(f"  Latest entry: {last_ts}")
    print()

    table = Table(show_lines=True)
    table.add_column("Entry ID", style="bold")
    table.add_column("Submitter")
    table.add_column("Timestamp")
    table.add_column("Hash")

    for entry in ledger["entries"]:
        table.add_row(
            entry["id"],
            entry["submitter"],
            entry["timestamp"],
            entry["hash"],
        )

    console.print(table)


def export_ledger_to_excel(ledger):
    """Export the full ledger to an Excel file."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Ledger"

    headers = ["Entry ID", "Submitter", "Timestamp", "Txn ID", "Hash"]
    ws.append(headers)

    for entry in ledger["entries"]:
        ws.append([
            entry["id"],
            entry["submitter"],
            entry["timestamp"],
            entry["transaction_id"],
            entry["hash"],
        ])

    export_path = os.path.join(os.path.dirname(LEDGER_FILE), "ledger_export.xlsx")
    wb.save(export_path)
    print(f"Ledger exported to {export_path}")


def import_transactions_from_excel(user):
    """Bulk-import transactions from an Excel file (.xlsx)."""
    print("Upload Transactions from Excel")
    print("Expected columns: Transaction ID, Date, Product ID, Amount, "
          "Currency, Quantity, Payer ID, Seller ID, On Credit")
    print()

    filepath = input("  Path to Excel file: ").strip().strip('"')

    if not filepath.lower().endswith(".xlsx"):
        print("File must be a .xlsx Excel file.")
        return
    if not os.path.isfile(filepath):
        print(f"File not found: {filepath}")
        return

    try:
        wb = load_workbook(filepath, data_only=True)
    except Exception as e:
        print(f"Could not open file: {e}")
        return

    ws = wb.active
    rows = list(ws.iter_rows(min_row=2, values_only=True))

    if not rows:
        print("No data rows found (only a header row or empty sheet).")
        return

    # Validators mapped to each field in FIELD_ORDER
    validators = {
        "transaction_id": validate_non_empty,
        "date": validate_date,
        "product_id": validate_non_empty,
        "amount": validate_amount,
        "currency": validate_non_empty,
        "quantity": validate_quantity,
        "payer_id": validate_non_empty,
        "seller_id": validate_non_empty,
        "on_credit": validate_yes_no,
    }

    valid_entries = []
    errors = []

    for row_idx, row in enumerate(rows, start=2):
        if len(row) < 9:
            errors.append((row_idx, "Row has fewer than 9 columns"))
            continue

        # Map cells to field names, converting to string
        fields = {}
        row_valid = True
        for col_idx, key in enumerate(FIELD_ORDER):
            cell_value = row[col_idx]
            raw = str(cell_value).strip() if cell_value is not None else ""
            result = validators[key](raw)
            if result is None:
                errors.append((row_idx, f"Invalid {key}: '{raw}'"))
                row_valid = False
                break
            fields[key] = result

        if not row_valid:
            continue

        canonical, normalized = normalize_transaction(fields)
        hash_value = hash_transaction(canonical)

        valid_entries.append({
            "transaction_id": normalized["transaction_id"],
            "hash": hash_value,
            "submitter": user["entity_name"],
            "timestamp": datetime.utcnow().strftime("%Y-%m-%dT%H:%M:%S"),
        })

    # Display summary
    print()
    print("Import Summary")
    print(f"  Rows processed: {len(rows)}")
    print(f"  Succeeded: {len(valid_entries)}")
    print(f"  Skipped: {len(errors)}")

    if errors:
        print()
        err_table = Table(title="Skipped Rows", show_lines=True)
        err_table.add_column("Row", style="bold")
        err_table.add_column("Reason")
        for row_num, reason in errors:
            err_table.add_row(str(row_num), reason)
        console.print(err_table)

    if not valid_entries:
        print("No valid transactions to submit.")
        return

    print()
    submit = ask_choice(
        f"Submit {len(valid_entries)} transaction(s) to the ledger?",
        ["y", "n"], default="y",
    )

    if submit == "y":
        ledger = load_ledger()
        for entry_data in valid_entries:
            entry_data["id"] = get_next_id(ledger)
            ledger["entries"].append(entry_data)
        save_ledger(ledger)
        print(f"{len(valid_entries)} transaction(s) submitted to the ledger.")
    else:
        print("Import cancelled.")


def search_by_hash(ledger):
    """Search the ledger by hash and check for counterparty verification."""
    hash_query = input("  Enter Hash to search: ").strip().lower()
    found = [e for e in ledger["entries"] if e["hash"].lower() == hash_query]

    if not found:
        print("No transactions found matching that hash.")
        return

    for entry in found:
        print("-" * 60)
        print("Transaction Found")
        print(f"  Txn ID:    {entry['transaction_id']}")
        print(f"  Submitter: {entry['submitter']}")
        print(f"  Timestamp: {entry['timestamp']}")
        print(f"  Hash:      {entry['hash']}")
        print()

        # Look for counterparty entry with the same transaction ID
        counterparty = [
            e for e in ledger["entries"]
            if e["transaction_id"] == entry["transaction_id"] and e["id"] != entry["id"]
        ]

        if not counterparty:
            print("  -- No counterparty submission yet --")
        else:
            for cp in counterparty:
                if cp["hash"] == entry["hash"]:
                    print("  -- Counterparty Hash Matches -- Transaction Verified --")
                else:
                    print("  -- Counterparty Hash Differs -- Transaction Flagged --")
                    print(f"  Counterparty Submitter: {cp['submitter']}")
                    print(f"  Counterparty Timestamp: {cp['timestamp']}")
                    print(f"  Counterparty Hash:      {cp['hash']}")
        print()


def search_by_entity(ledger):
    """Search the ledger by entity name."""
    entity = input("  Enter Entity Name to search: ").strip().upper().replace(" ", "")
    found = [e for e in ledger["entries"] if e["submitter"] == entity]

    if not found:
        print(f"No blocks found for entity: {entity}")
        return

    display_chain_stats(ledger)

    print(f"Found {len(found)} transaction(s) submitted by {entity}:")
    print()
    table = Table(title=f"Block #1 -- {entity} Transactions", show_lines=True)
    table.add_column("Entry ID", style="bold")
    table.add_column("Submitter")
    table.add_column("Timestamp")
    table.add_column("Hash")

    for entry in found:
        table.add_row(
            entry["id"],
            entry["submitter"],
            entry["timestamp"],
            entry["hash"],
        )

    console.print(table)


def auditor_flow(user):
    """Main workflow for the auditor role."""
    while True:
        clear_screen()
        display_banner()
        print(f"Logged in as: {user['entity_name']} (Auditor)")
        print()
        print("Auditor Menu")
        print("  [1] View Full Ledger")
        print("  [2] View Verification Status")
        print("  [3] Search by Hash")
        print("  [4] Search by Entity")
        print("  [5] Logout")
        print()

        choice = ask_choice("  Select an option", ["1", "2", "3", "4", "5"])
        print()

        ledger = load_ledger()

        if choice == "1":
            display_full_ledger(ledger)
            if ledger["entries"]:
                export = ask_choice("Export ledger to Excel?", ["y", "n"], default="n")
                if export == "y":
                    export_ledger_to_excel(ledger)
        elif choice == "2":
            display_verification_status(ledger)
        elif choice == "3":
            search_by_hash(ledger)
        elif choice == "4":
            search_by_entity(ledger)
        elif choice == "5":
            break

        print()
        input("Press Enter to continue...")


# --- Main Menu and Login ---


def main_menu(user):
    """Show the appropriate menu based on user role."""
    if user["role"] == "auditor":
        auditor_flow(user)
    else:
        while True:
            clear_screen()
            display_banner()
            print(f"Logged in as: {user['entity_name']}")
            print()
            print("Company Menu")
            print("  [1] Enter Transaction")
            print("  [2] Upload Transactions from Excel")
            print("  [3] View My Submissions")
            print("  [4] Logout")
            print()

            choice = ask_choice("  Select an option", ["1", "2", "3", "4"])

            if choice == "1":
                company_flow(user)
            elif choice == "2":
                clear_screen()
                display_banner()
                import_transactions_from_excel(user)
                print()
                input("Press Enter to continue...")
            elif choice == "3":
                clear_screen()
                display_banner()
                ledger = load_ledger()
                found = [e for e in ledger["entries"] if e["submitter"] == user["entity_name"]]
                if not found:
                    print("You have no submissions yet.")
                else:
                    table = Table(title=f"Submissions by {user['entity_name']}", show_lines=True)
                    table.add_column("Entry ID", style="bold")
                    table.add_column("Txn ID")
                    table.add_column("Timestamp")
                    table.add_column("Hash")

                    for entry in found:
                        table.add_row(
                            entry["id"],
                            entry["transaction_id"],
                            entry["timestamp"],
                            entry["hash"],
                        )
                    console.print(table)

                print()
                input("Press Enter to continue...")
            elif choice == "4":
                break


def login():
    """Authenticate via public key. Returns user dict, 'quit', or None on failure."""
    for attempt in range(3):
        key = input("  Public Key: ").strip().upper()
        if key in ("QUIT", "EXIT", "Q"):
            return "quit"
        if key == AUDITOR_KEY:
            print(f"\nAuditor view accessed.")
            return {"role": "auditor", "entity_name": "AUDITOR", "username": key}
        if key in COMPANY_KEYS:
            user = COMPANY_KEYS[key].copy()
            user["username"] = key
            print(f"\nAuthenticated with public key: {key}")
            return user
        remaining = 2 - attempt
        if remaining > 0:
            print(f"  Unknown public key. {remaining} attempt(s) remaining.")
        else:
            print("  Too many failed attempts.")
    return None


def main():
    """Entry point: login loop and routing."""
    try:
        while True:
            clear_screen()
            display_banner()
            print("Public Key Login")
            print("Auditor view: VIEW")
            print("Type 'quit' to exit.")
            print()

            user = login()
            if user is None:
                print()
                retry = ask_choice("Try again?", ["y", "n"], default="y")
                if retry != "y":
                    break
            elif user == "quit":
                break
            else:
                main_menu(user)

        print("\nGoodbye.")
    except KeyboardInterrupt:
        print("\n\nSession ended.")


if __name__ == "__main__":
    main()
