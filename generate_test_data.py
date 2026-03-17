"""
generate_test_data.py
---------------------
Generates two Excel test data files for the Blockchain Audit Verification System:
  - company_a_transactions.xlsx  (Company A / payer formatting style)
  - vendor_b_transactions.xlsx   (Vendor B / seller formatting style)

Both files contain the same 100 underlying transactions but use different
real-world formatting conventions to demonstrate that the normalization engine
produces identical hashes regardless of input format.

Run:  python generate_test_data.py
"""

import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import date, timedelta

# ---------------------------------------------------------------------------
# Canonical transaction data (100 transactions, all shared values)
# ---------------------------------------------------------------------------

PRODUCTS = [
    ("WIDGET-500",  25.00),   # unit price
    ("GADGET-100",  149.99),
    ("PART-200",    8.50),
    ("CABLE-300",   12.75),
    ("SENSOR-400",  340.00),
    ("BRACKET-600", 4.25),
    ("MODULE-700",  875.00),
    ("CHIP-800",    62.00),
]

# Spread 100 transaction dates across 2025 business days
def business_days_2025(n: int) -> list[date]:
    """Return n evenly-spaced business days across 2025."""
    start = date(2025, 1, 6)
    end   = date(2025, 12, 29)
    total_days = (end - start).days
    step = total_days // (n - 1)
    days = []
    d = start
    count = 0
    while count < n:
        if d.weekday() < 5:  # Monday–Friday
            days.append(d)
            count += 1
        d += timedelta(days=1)
        if d > end and count < n:
            d = start  # wrap (shouldn't happen with these params)
    return days[:n]

dates = business_days_2025(100)

# Build canonical list of 100 transactions
TRANSACTIONS = []
for i in range(100):
    idx = i + 1
    product, unit_price = PRODUCTS[i % len(PRODUCTS)]
    quantity = 10 + (i * 7 % 491)        # 10–500, varied
    amount = round(unit_price * quantity, 2)
    on_credit_bool = (i % 5 == 0) or (i % 7 == 0)  # ~37% YES

    TRANSACTIONS.append({
        "txn_id":    f"INV-{idx:03d}",
        "date":      dates[i].strftime("%Y-%m-%d"),
        "product_id": product,
        "amount":    amount,
        "currency":  "USD",
        "quantity":  float(quantity),
        "payer_id":  "COMPANYA",
        "seller_id": "VENDORB",
        "on_credit": on_credit_bool,
    })

# ---------------------------------------------------------------------------
# Formatting helpers
# ---------------------------------------------------------------------------

def fmt_amount_a(val: float) -> str:
    """Company A: accounting format  $10,000.00"""
    return f"${val:,.2f}"

def fmt_amount_b(val: float) -> str:
    """Vendor B: plain decimal  10000.00"""
    return f"{val:.2f}"

def fmt_qty_a(val: float) -> str:
    """Company A: whole integer  100"""
    return str(int(val))

def fmt_qty_b(val: float) -> str:
    """Vendor B: decimal notation  100.00"""
    return f"{val:.2f}"

def fmt_credit_a(val: bool) -> str:
    """Company A: lowercase text  yes / no"""
    return "yes" if val else "no"

def fmt_credit_b(val: bool) -> str:
    """Vendor B: numeric boolean  1 / 0"""
    return "1" if val else "0"

# ---------------------------------------------------------------------------
# Excel writing helpers
# ---------------------------------------------------------------------------

HEADERS = [
    "Transaction ID",
    "Date (UTC)",
    "Product ID",
    "Total Amount",
    "Currency",
    "Quantity",
    "Payer ID",
    "Seller ID",
    "On Credit",
]

def style_header_row(ws, header_fill_hex: str):
    header_fill = PatternFill("solid", fgColor=header_fill_hex)
    header_font = Font(bold=True, color="FFFFFF")
    for col, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

def set_column_widths(ws):
    widths = [16, 12, 14, 16, 10, 10, 14, 12, 12]
    for col, width in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width

# ---------------------------------------------------------------------------
# Write Company A file
# ---------------------------------------------------------------------------

def write_company_a(path: str):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Transactions"

    style_header_row(ws, "003DA5")  # Blue

    for row_idx, txn in enumerate(TRANSACTIONS, start=2):
        ws.cell(row=row_idx, column=1, value=txn["txn_id"].upper())          # INV-001
        ws.cell(row=row_idx, column=2, value=txn["date"])                    # 2025-03-15
        ws.cell(row=row_idx, column=3, value=txn["product_id"].upper())      # WIDGET-500
        ws.cell(row=row_idx, column=4, value=fmt_amount_a(txn["amount"]))    # $10,000.00
        ws.cell(row=row_idx, column=5, value=txn["currency"].upper())        # USD
        ws.cell(row=row_idx, column=6, value=fmt_qty_a(txn["quantity"]))     # 100
        ws.cell(row=row_idx, column=7, value=txn["payer_id"].upper())        # DEPAUL
        ws.cell(row=row_idx, column=8, value=txn["seller_id"].upper())       # VENDORB
        ws.cell(row=row_idx, column=9, value=fmt_credit_a(txn["on_credit"])) # yes / no

        # Alternate row shading
        if row_idx % 2 == 0:
            fill = PatternFill("solid", fgColor="EAF2FB")
            for col in range(1, 10):
                ws.cell(row=row_idx, column=col).fill = fill

    set_column_widths(ws)
    ws.freeze_panes = "A2"
    wb.save(path)
    print(f"[OK] Written: {path}")

# ---------------------------------------------------------------------------
# Write Vendor B file
# ---------------------------------------------------------------------------

def write_vendor_b(path: str):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Transactions"

    style_header_row(ws, "2D6A4F")  # deep green — vendor feel

    for row_idx, txn in enumerate(TRANSACTIONS, start=2):
        ws.cell(row=row_idx, column=1, value=txn["txn_id"].lower())          # inv-001
        ws.cell(row=row_idx, column=2, value=txn["date"])                    # 2025-03-15
        ws.cell(row=row_idx, column=3, value=txn["product_id"].lower())      # widget-500
        ws.cell(row=row_idx, column=4, value=fmt_amount_b(txn["amount"]))    # 10000.00
        ws.cell(row=row_idx, column=5, value=txn["currency"].lower())        # usd
        ws.cell(row=row_idx, column=6, value=fmt_qty_b(txn["quantity"]))     # 100.00
        ws.cell(row=row_idx, column=7, value=txn["payer_id"].lower())        # depaul
        ws.cell(row=row_idx, column=8, value=txn["seller_id"].lower())       # vendorb
        ws.cell(row=row_idx, column=9, value=fmt_credit_b(txn["on_credit"])) # 1 / 0

        # Alternate row shading
        if row_idx % 2 == 0:
            fill = PatternFill("solid", fgColor="E9F5EE")
            for col in range(1, 10):
                ws.cell(row=row_idx, column=col).fill = fill

    set_column_widths(ws)
    ws.freeze_panes = "A2"
    wb.save(path)
    print(f"[OK] Written: {path}")

# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    out_dir = "Test Data"
    os.makedirs(out_dir, exist_ok=True)
    write_company_a(os.path.join(out_dir, "company_a_transactions.xlsx"))
    write_vendor_b(os.path.join(out_dir, "vendor_b_transactions.xlsx"))

    print()
    print("Formatting differences (same underlying data, different styles):")
    print("  Amount    : Company A uses $X,XXX.XX  |  Vendor B uses plain XXXXX.XX")
    print("  Case      : Company A uses UPPERCASE  |  Vendor B uses lowercase")
    print("  Quantity  : Company A uses integers   |  Vendor B uses X.XX decimals")
    print("  On Credit : Company A uses yes/no     |  Vendor B uses 1/0")
    print()
    print("Both files should produce VERIFIED (matching hashes) when imported.")
